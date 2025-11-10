"""
Test suite for excel_export.py module

This test suite validates the Excel export functionality with realistic data
and ensures proper formatting, error handling, and file generation.
"""

import unittest
import tempfile
import os
from pathlib import Path
from typing import List, Tuple
from openpyxl import load_workbook

# Import module under test
from excel_export import (
    export_solutions_to_excel,
    export_solutions_to_bytes,
    _calculate_weighted_averages
)
from inventory import LotData


class TestExcelExport(unittest.TestCase):
    """Test cases for Excel export functionality"""

    def setUp(self):
        """Set up test fixtures"""
        # Create sample LotData objects
        self.lot1 = LotData(
            article_code='3|POB',
            lot_code='TEST-LOT-001',
            description='Test lot 1',
            dc_real=85.0,
            fp_real=750.0,
            duck_real=10.0,
            other_elements_real=5.0,
            feather_real=10.0,
            qty_available=1000.0,
            cost_per_kg=100.0,
            lab_notes='Test notes for lot 1',
            is_estimated=False
        )

        self.lot2 = LotData(
            article_code='3|POB',
            lot_code='TEST-LOT-002',
            description='Test lot 2',
            dc_real=80.0,
            fp_real=700.0,
            duck_real=15.0,
            other_elements_real=7.0,
            feather_real=13.0,
            qty_available=2000.0,
            cost_per_kg=90.0,
            lab_notes='Test notes for lot 2',
            is_estimated=False
        )

        self.lot3 = LotData(
            article_code='3|POBPW',
            lot_code='TEST-LOT-003',
            description='Test lot 3',
            dc_real=90.0,
            fp_real=800.0,
            duck_real=5.0,
            other_elements_real=3.0,
            feather_real=7.0,
            qty_available=500.0,
            cost_per_kg=120.0,
            lab_notes='Test notes for lot 3',
            is_estimated=False
        )

        # Create sample solutions
        self.solutions = [
            # Solution 1: 2 lots
            (
                [self.lot1, self.lot2],
                [600.0, 1400.0],
                1234.56
            ),
            # Solution 2: 3 lots
            (
                [self.lot1, self.lot2, self.lot3],
                [500.0, 1000.0, 500.0],
                1150.25
            )
        ]

        # Sample requirements
        self.requirements = {
            'dc': 80.0,
            'fp': 750.0,
            'duck': 10.0,
            'qty': 2000.0
        }

    def test_export_to_file_success(self):
        """Test successful export to file"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # Export solutions
            export_solutions_to_excel(
                solutions=self.solutions,
                requirements=self.requirements,
                output_path=tmp_path
            )

            # Verify file was created
            self.assertTrue(os.path.exists(tmp_path), "Excel file should be created")
            self.assertGreater(os.path.getsize(tmp_path), 0, "Excel file should not be empty")

            # Load and verify content
            wb = load_workbook(tmp_path)
            self.assertIn('TUTTE_LE_SOLUZIONI', wb.sheetnames, "Sheet name should be correct")

            ws = wb['TUTTE_LE_SOLUZIONI']
            self.assertGreater(ws.max_row, 0, "Worksheet should have content")

            # Verify solution headers exist
            found_solution_headers = 0
            for row in ws.iter_rows(values_only=True):
                if row[0] and 'SOLUZIONE' in str(row[0]):
                    found_solution_headers += 1

            self.assertEqual(found_solution_headers, 2, "Should have 2 solution headers")

        finally:
            # Clean up
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_export_to_bytes_success(self):
        """Test successful export to bytes (in-memory)"""
        excel_bytes = export_solutions_to_bytes(
            solutions=self.solutions,
            requirements=self.requirements
        )

        self.assertIsInstance(excel_bytes, bytes, "Should return bytes")
        self.assertGreater(len(excel_bytes), 0, "Bytes should not be empty")

        # Verify it's a valid Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp.write(excel_bytes)
            tmp_path = tmp.name

        try:
            wb = load_workbook(tmp_path)
            self.assertIn('TUTTE_LE_SOLUZIONI', wb.sheetnames)
        finally:
            os.unlink(tmp_path)

    def test_export_empty_solutions_raises_error(self):
        """Test that empty solutions list raises ValueError"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name

        try:
            with self.assertRaises(ValueError) as context:
                export_solutions_to_excel(
                    solutions=[],
                    requirements=self.requirements,
                    output_path=tmp_path
                )
            self.assertIn('empty', str(context.exception).lower())
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_export_invalid_path_raises_error(self):
        """Test that invalid output path raises ValueError"""
        with self.assertRaises(ValueError):
            export_solutions_to_excel(
                solutions=self.solutions,
                requirements=self.requirements,
                output_path=''
            )

    def test_calculate_weighted_averages(self):
        """Test weighted average calculation"""
        combination = [self.lot1, self.lot2]
        allocations = [600.0, 1400.0]

        metrics = _calculate_weighted_averages(combination, allocations)

        # Verify structure
        self.assertIn('total_kg', metrics)
        self.assertIn('dc_avg', metrics)
        self.assertIn('duck_avg', metrics)
        self.assertIn('fp_avg', metrics)
        self.assertIn('cost_per_kg', metrics)
        self.assertIn('lot_count', metrics)

        # Verify values
        self.assertEqual(metrics['total_kg'], 2000.0)
        self.assertEqual(metrics['lot_count'], 2)

        # Verify weighted averages
        # DC: (85*600 + 80*1400) / 2000 = 81.5
        expected_dc = (85.0 * 600.0 + 80.0 * 1400.0) / 2000.0
        self.assertAlmostEqual(metrics['dc_avg'], expected_dc, places=2)

        # FP: (750*600 + 700*1400) / 2000 = 715
        expected_fp = (750.0 * 600.0 + 700.0 * 1400.0) / 2000.0
        self.assertAlmostEqual(metrics['fp_avg'], expected_fp, places=2)

        # Cost: (100*600 + 90*1400) / 2000 = 93
        expected_cost = (100.0 * 600.0 + 90.0 * 1400.0) / 2000.0
        self.assertAlmostEqual(metrics['cost_per_kg'], expected_cost, places=2)

    def test_mismatched_lots_and_allocations(self):
        """Test handling of mismatched lots and allocations"""
        # Create mismatched solution
        bad_solutions = [
            (
                [self.lot1, self.lot2],  # 2 lots
                [1000.0],  # Only 1 allocation (mismatch!)
                100.0
            )
        ]

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # Should not crash, but skip the bad solution
            export_solutions_to_excel(
                solutions=bad_solutions,
                requirements=self.requirements,
                output_path=tmp_path
            )

            # File should still be created (but mostly empty)
            self.assertTrue(os.path.exists(tmp_path))

        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)

    def test_lot_with_none_values(self):
        """Test handling of lots with None values"""
        # Create lot with many None values
        sparse_lot = LotData(
            article_code='3|POB',
            lot_code='SPARSE-LOT',
            description='Lot with missing data',
            dc_real=85.0,  # Only DC is present
            fp_real=None,
            duck_real=None,
            other_elements_real=None,
            feather_real=None,
            qty_available=1000.0,
            cost_per_kg=None,
            lab_notes='',
            is_estimated=True
        )

        sparse_solutions = [
            ([sparse_lot], [1000.0], 500.0)
        ]

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name

        try:
            # Should handle None values gracefully
            export_solutions_to_excel(
                solutions=sparse_solutions,
                requirements=self.requirements,
                output_path=tmp_path
            )

            self.assertTrue(os.path.exists(tmp_path))

            # Verify file can be loaded
            wb = load_workbook(tmp_path)
            ws = wb['TUTTE_LE_SOLUZIONI']
            self.assertGreater(ws.max_row, 0)

        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)


class TestExcelExportIntegration(unittest.TestCase):
    """Integration tests with real-like data"""

    def test_export_matches_example_format(self):
        """Test that export matches the example file format"""
        # This test would compare structure with esempio_output_excel.xlsx
        # For now, we verify key structural elements

        lot1 = LotData(
            article_code='3|POB',
            lot_code='25/03/05-SAN-WGD80',
            description='Test description',
            dc_real=77.3,
            fp_real=627.0,
            duck_real=20.13,
            other_elements_real=7.4,
            feather_real=10.6,
            qty_available=4100.0,
            cost_per_kg=93.8238,
            lab_notes='Test lab notes',
            is_estimated=False,
            dc_nominal=80.0,
            quality_nominal='CIN',
            fp_nominal=0.0
        )

        solutions = [([lot1], [3763.34], 1574.07)]
        requirements = {'dc': 80.0, 'qty': 5000.0}

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name

        try:
            export_solutions_to_excel(solutions, requirements, tmp_path)

            # Load and verify structure
            wb = load_workbook(tmp_path)
            ws = wb['TUTTE_LE_SOLUZIONI']

            # Row 1 should be solution header
            header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            self.assertIn('SOLUZIONE', str(header_row[0]))
            self.assertIn('Score', str(header_row[0]))

            # Row 2 should be column headers
            column_headers = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
            self.assertEqual(column_headers[0], 'Codice Art')
            self.assertEqual(column_headers[5], 'Codice Lotto')

            # Row 3 should be data
            data_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
            self.assertEqual(data_row[0], '3|POB')
            self.assertEqual(data_row[5], '25/03/05-SAN-WGD80')

        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)


def run_tests():
    """Run all tests"""
    unittest.main(argv=[''], verbosity=2, exit=False)


if __name__ == '__main__':
    print("=" * 60)
    print("Excel Export Module - Test Suite")
    print("=" * 60)
    print()
    run_tests()
