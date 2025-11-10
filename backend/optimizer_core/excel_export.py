"""
OPTIMIZER v3.3 - Excel Export Module
Exports optimization solutions to professionally formatted Excel files

This module generates Excel reports from optimizer results, formatting them
according to the standard template with proper styling, summaries, and metrics.
"""

import logging
from typing import List, Tuple, Dict, Optional, Any
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from inventory import LotData
from config import OUTPUT_EXCEL_COLUMNS, EXCEL_COLORS

# Configure logging
logger = logging.getLogger(__name__)


def export_solutions_to_excel(
    solutions: List[Tuple[List[LotData], List[float], float]],
    requirements: Dict[str, Any],
    output_path: str
) -> None:
    """
    Export optimization solutions to Excel file.

    Creates a professionally formatted Excel workbook containing all optimization
    solutions with detailed lot information, metrics, and summary statistics.

    Args:
        solutions: List of solution tuples, where each tuple contains:
            - combination (List[LotData]): List of lots in the solution
            - allocations (List[float]): Kg allocated to each lot
            - score (float): Optimization score for the solution
        requirements: Dictionary with optimization requirements (dc, fp, duck, qty, etc.)
        output_path: File path where Excel file will be saved

    Raises:
        ValueError: If solutions is empty or invalid
        IOError: If file cannot be written
        Exception: For other unexpected errors during export

    Example:
        >>> solutions = [(
        ...     [lot1, lot2],  # combination
        ...     [500.0, 1500.0],  # allocations
        ...     1234.56  # score
        ... )]
        >>> requirements = {'dc': 80.0, 'fp': 750.0, 'qty': 2000.0}
        >>> export_solutions_to_excel(solutions, requirements, '/tmp/output.xlsx')
    """
    logger.info(f"Starting Excel export to: {output_path}")
    logger.info(f"Number of solutions to export: {len(solutions)}")

    # Validate inputs
    if not solutions:
        logger.error("Cannot export: solutions list is empty")
        raise ValueError("Solutions list cannot be empty")

    if not output_path:
        logger.error("Cannot export: output_path is empty")
        raise ValueError("Output path must be specified")

    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "TUTTE_LE_SOLUZIONI"

        logger.debug(f"Created workbook with sheet: {ws.title}")

        # Track current row position
        current_row = 1

        # Process each solution
        for solution_idx, (combination, allocations, score) in enumerate(solutions, 1):
            logger.debug(f"Processing solution {solution_idx}/{len(solutions)}")

            # Validate solution data
            if len(combination) != len(allocations):
                logger.warning(
                    f"Solution {solution_idx}: Mismatch between lots ({len(combination)}) "
                    f"and allocations ({len(allocations)}). Skipping."
                )
                continue

            # Write solution header
            current_row = _write_solution_header(
                ws, current_row, solution_idx, score
            )

            # Write column headers
            current_row = _write_column_headers(ws, current_row)

            # Write lot data rows
            current_row = _write_lot_data(
                ws, current_row, combination, allocations
            )

            # Calculate and write summary row
            current_row = _write_solution_summary(
                ws, current_row, combination, allocations
            )

            # Add spacing between solutions
            current_row += 2

        # Auto-size columns
        _auto_size_columns(ws)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Successfully exported {len(solutions)} solutions to {output_path}")

    except PermissionError as e:
        logger.error(f"Permission denied writing to {output_path}: {e}")
        raise IOError(f"Cannot write to {output_path}: Permission denied") from e
    except Exception as e:
        logger.error(f"Unexpected error during Excel export: {e}", exc_info=True)
        raise


def _write_solution_header(
    ws,
    start_row: int,
    solution_number: int,
    score: float
) -> int:
    """
    Write solution separator header.

    Args:
        ws: Worksheet object
        start_row: Starting row number
        solution_number: Solution index (1-based)
        score: Optimization score

    Returns:
        Next available row number
    """
    header_text = f"═══ SOLUZIONE {solution_number} - Score: {score:.2f} ═══"
    cell = ws.cell(row=start_row, column=1)
    cell.value = header_text

    # Style: Bold, larger font, blue background
    cell.font = Font(bold=True, size=12)
    cell.fill = PatternFill(start_color=EXCEL_COLORS['header'], end_color=EXCEL_COLORS['header'], fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')

    logger.debug(f"Wrote header for solution {solution_number} at row {start_row}")
    return start_row + 1


def _write_column_headers(ws, start_row: int) -> int:
    """
    Write column headers for lot data table.

    Args:
        ws: Worksheet object
        start_row: Starting row number

    Returns:
        Next available row number
    """
    for col_idx, column_name in enumerate(OUTPUT_EXCEL_COLUMNS, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = column_name

        # Style: Bold, white text on blue background, centered
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color=EXCEL_COLORS['header'], end_color=EXCEL_COLORS['header'], fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Add borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        cell.border = thin_border

    logger.debug(f"Wrote column headers at row {start_row}")
    return start_row + 1


def _write_lot_data(
    ws,
    start_row: int,
    combination: List[LotData],
    allocations: List[float]
) -> int:
    """
    Write lot data rows for a solution.

    Args:
        ws: Worksheet object
        start_row: Starting row number
        combination: List of LotData objects
        allocations: List of kg allocated to each lot

    Returns:
        Next available row number
    """
    current_row = start_row

    # Calculate total kg for percentage calculation
    total_kg = sum(allocations)

    for lot, kg_used in zip(combination, allocations):
        # Get lot data as dictionary
        lot_dict = lot.to_dict()

        # Add allocation columns
        lot_dict['Kg tot usati in miscela'] = kg_used
        lot_dict['% di miscela'] = (kg_used / total_kg * 100) if total_kg > 0 else 0

        # Write each column
        for col_idx, column_name in enumerate(OUTPUT_EXCEL_COLUMNS, 1):
            cell = ws.cell(row=current_row, column=col_idx)
            value = lot_dict.get(column_name)

            # Handle None, NaN, and 'nan' string values
            if value is None or (isinstance(value, float) and str(value).lower() == 'nan') or (isinstance(value, str) and value.lower() == 'nan'):
                cell.value = None
            else:
                cell.value = value

            # Apply formatting based on column type
            _format_cell(cell, column_name, value, lot)

        current_row += 1

    logger.debug(f"Wrote {len(combination)} lot data rows starting at row {start_row}")
    return current_row


def _format_cell(cell, column_name: str, value: Any, lot: LotData) -> None:
    """
    Apply formatting to individual cell based on column type and value.

    Args:
        cell: Cell object to format
        column_name: Name of the column
        value: Cell value
        lot: LotData object for the row
    """
    # Borders for all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.border = thin_border
    cell.alignment = Alignment(vertical='center')

    # Number formatting for specific columns
    if 'reale' in column_name or 'Nominale' in column_name or '% di miscela' in column_name:
        if value is not None and value != '':
            cell.number_format = '0.00'
    elif 'Costo' in column_name or 'euro/kg' in column_name:
        if value is not None and value != '':
            cell.number_format = '€#,##0.00'
    elif 'Quantità' in column_name or 'Kg tot usati' in column_name:
        if value is not None and value != '':
            cell.number_format = '#,##0.00'

    # Highlight estimated data
    if column_name == 'Stimato si/no' and value == 'SI':
        cell.fill = PatternFill(start_color=EXCEL_COLORS['estimated'], end_color=EXCEL_COLORS['estimated'], fill_type='solid')
        cell.font = Font(bold=True)


def _write_solution_summary(
    ws,
    start_row: int,
    combination: List[LotData],
    allocations: List[float]
) -> int:
    """
    Write summary row with total metrics for the solution.

    Args:
        ws: Worksheet object
        start_row: Starting row number
        combination: List of LotData objects
        allocations: List of kg allocated to each lot

    Returns:
        Next available row number
    """
    # Calculate metrics
    metrics = _calculate_weighted_averages(combination, allocations)

    # Format summary text
    summary_text = (
        f"📊 Totale: {metrics['total_kg']:.2f} kg | "
        f"DC: {metrics['dc_avg']:.2f}% | "
        f"Duck: {metrics['duck_avg']:.2f}% | "
        f"FP: {metrics['fp_avg']:.2f} | "
        f"OE: {metrics['oe_avg']:.2f}% | "
        f"Costo: €{metrics['cost_per_kg']:.2f}/kg | "
        f"Lotti: {metrics['lot_count']}"
    )

    # Write summary in first column
    cell = ws.cell(row=start_row, column=1)
    cell.value = summary_text

    # Style: Bold, green background
    cell.font = Font(bold=True, size=11)
    cell.fill = PatternFill(start_color=EXCEL_COLORS['optimal'], end_color=EXCEL_COLORS['optimal'], fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')

    logger.debug(f"Wrote summary at row {start_row}: {summary_text}")
    return start_row + 1


def _calculate_weighted_averages(
    combination: List[LotData],
    allocations: List[float]
) -> Dict[str, float]:
    """
    Calculate weighted averages for solution metrics.

    Args:
        combination: List of LotData objects
        allocations: List of kg allocated to each lot

    Returns:
        Dictionary with calculated metrics:
        - total_kg: Total kilograms in the blend
        - dc_avg: Weighted average Down Cluster %
        - duck_avg: Weighted average Duck %
        - fp_avg: Weighted average Fill Power
        - oe_avg: Weighted average Other Elements %
        - cost_per_kg: Weighted average cost per kg
        - lot_count: Number of lots in the blend
    """
    total_kg = sum(allocations)

    if total_kg == 0:
        logger.warning("Total kg is zero, returning empty metrics")
        return {
            'total_kg': 0,
            'dc_avg': 0,
            'duck_avg': 0,
            'fp_avg': 0,
            'oe_avg': 0,
            'cost_per_kg': 0,
            'lot_count': 0
        }

    # Calculate weighted averages
    dc_avg = 0
    duck_avg = 0
    fp_avg = 0
    oe_avg = 0
    total_cost = 0

    for lot, kg in zip(combination, allocations):
        weight = kg / total_kg

        # DC average
        if lot.dc_real is not None:
            dc_avg += lot.dc_real * weight

        # Duck average
        if lot.duck_real is not None:
            duck_avg += lot.duck_real * weight

        # FP average
        if lot.fp_real is not None:
            fp_avg += lot.fp_real * weight

        # OE average (Other Elements)
        if lot.other_elements_real is not None:
            oe_avg += lot.other_elements_real * weight

        # Cost
        if lot.cost_per_kg is not None:
            total_cost += lot.cost_per_kg * kg

    cost_per_kg = total_cost / total_kg if total_kg > 0 else 0

    metrics = {
        'total_kg': total_kg,
        'dc_avg': dc_avg,
        'duck_avg': duck_avg,
        'fp_avg': fp_avg,
        'oe_avg': oe_avg,
        'cost_per_kg': cost_per_kg,
        'lot_count': len(combination)
    }

    logger.debug(f"Calculated metrics: {metrics}")
    return metrics


def _auto_size_columns(ws) -> None:
    """
    Auto-size columns based on content width.

    Args:
        ws: Worksheet object
    """
    for column_cells in ws.columns:
        max_length = 0
        column_letter = None

        for cell in column_cells:
            if cell.value:
                # Get column letter from first cell
                if column_letter is None:
                    column_letter = cell.column_letter

                # Calculate content length
                try:
                    cell_length = len(str(cell.value))
                    max_length = max(max_length, cell_length)
                except:
                    pass

        # Set column width with some padding
        if column_letter:
            # Limit max width to avoid extremely wide columns
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    logger.debug("Auto-sized all columns")


def export_solutions_to_bytes(
    solutions: List[Tuple[List[LotData], List[float], float]],
    requirements: Dict[str, Any]
) -> bytes:
    """
    Export optimization solutions to Excel file as bytes (in-memory).

    This is useful for web applications that need to send Excel files
    without saving to disk first.

    Args:
        solutions: List of solution tuples (same format as export_solutions_to_excel)
        requirements: Dictionary with optimization requirements

    Returns:
        Excel file content as bytes

    Raises:
        ValueError: If solutions is empty or invalid
        Exception: For unexpected errors during export

    Example:
        >>> excel_bytes = export_solutions_to_bytes(solutions, requirements)
        >>> # Send bytes directly in HTTP response
    """
    from io import BytesIO

    logger.info(f"Starting in-memory Excel export for {len(solutions)} solutions")

    # Validate inputs
    if not solutions:
        logger.error("Cannot export: solutions list is empty")
        raise ValueError("Solutions list cannot be empty")

    try:
        # Create workbook (same logic as file export)
        wb = Workbook()
        ws = wb.active
        ws.title = "TUTTE_LE_SOLUZIONI"

        current_row = 1

        for solution_idx, (combination, allocations, score) in enumerate(solutions, 1):
            if len(combination) != len(allocations):
                logger.warning(f"Solution {solution_idx}: Lot/allocation mismatch, skipping")
                continue

            current_row = _write_solution_header(ws, current_row, solution_idx, score)
            current_row = _write_column_headers(ws, current_row)
            current_row = _write_lot_data(ws, current_row, combination, allocations)
            current_row = _write_solution_summary(ws, current_row, combination, allocations)
            current_row += 2

        _auto_size_columns(ws)

        # Save to BytesIO instead of file
        output = BytesIO()
        wb.save(output)
        excel_bytes = output.getvalue()

        logger.info(f"Successfully created in-memory Excel file ({len(excel_bytes)} bytes)")
        return excel_bytes

    except Exception as e:
        logger.error(f"Unexpected error during in-memory Excel export: {e}", exc_info=True)
        raise


if __name__ == '__main__':
    # Example usage / manual testing
    print("Excel Export Module - v3.3")
    print("This module is designed to be imported, not run directly.")
    print("For testing, use the optimizer module which calls this export function.")
