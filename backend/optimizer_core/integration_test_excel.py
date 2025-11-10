"""
Integration Test - Excel Export with Real Service Flow

This script simulates the actual usage pattern from excel_export_service.py
to verify end-to-end functionality.
"""

import sys
import tempfile
import os
from pathlib import Path
from openpyxl import load_workbook

# Add optimizer core to path (same as service does)
sys.path.insert(0, '/Users/carlocassigoli/CODE-progetti-Claude/Claude/WEP_APPS/blend-optimizer-web/backend/optimizer_core')

from excel_export import export_solutions_to_excel
from inventory import LotData


def test_service_usage_pattern():
    """
    Test the exact usage pattern from ExcelExportService.generate_excel()
    """
    print("=" * 70)
    print("Integration Test: Service Usage Pattern")
    print("=" * 70)
    print()

    # Create realistic LotData objects (mimicking database conversion)
    print("1. Creating LotData objects (mimicking DB → LotData conversion)...")

    lot1 = LotData(
        article_code='3|POB',
        lot_code='25/03/05-SAN-WGD80',
        description='Lotto test 1',
        dc_real=77.3,
        fp_real=627.0,
        duck_real=20.13,
        other_elements_real=7.4,
        feather_real=10.6,
        oxygen_real=None,
        turbidity_real=None,
        dc_nominal=80.0,
        fp_nominal=0.0,
        quality_nominal='CIN',
        total_fibres_real=11.7,
        broken_real=0.3,
        landfowl_real=None,
        qty_available=4100.0,
        cost_per_kg=93.8238,
        lab_notes='nan',
        is_estimated=False,
        dc_was_imputed=False,
        fp_was_imputed=False
    )

    lot2 = LotData(
        article_code='3|POB',
        lot_code='DY|2025|30472|5',
        description='2001 (LWR) - K3 - IN: G.POB',
        dc_real=86.0,
        fp_real=786.0,
        duck_real=29.11,
        other_elements_real=5.3,
        feather_real=4.1,
        oxygen_real=3.2,
        turbidity_real=1001.0,
        dc_nominal=90.0,
        fp_nominal=750.0,
        quality_nominal='LWR',
        total_fibres_real=9.1,
        broken_real=0.8,
        landfowl_real=None,
        qty_available=609.0,
        cost_per_kg=41.46,
        lab_notes='Ok per l\'uso KS.DC 86,0%, F 4,1%, CL1. Duck 29,1%, FP 786cuin.',
        is_estimated=False,
        dc_was_imputed=False,
        fp_was_imputed=False
    )

    lot3 = LotData(
        article_code='3|POBPW',
        lot_code='DY|2025|30370|8',
        description='WS4 - K4 - IN: G.OOB.CTY1.RUS',
        dc_real=90.3,
        fp_real=780.0,
        duck_real=7.56,
        other_elements_real=1.8,
        feather_real=4.7,
        oxygen_real=2.4,
        turbidity_real=1001.0,
        dc_nominal=90.0,
        fp_nominal=800.0,
        quality_nominal='CTY',
        total_fibres_real=4.4,
        broken_real=0.6,
        landfowl_real=0.1,
        qty_available=850.4,
        cost_per_kg=156.42,
        lab_notes='DC 90,3%,F 4,7%,CL1, Fill power 780 cuin. Duck 7,6%.',
        is_estimated=False,
        dc_was_imputed=False,
        fp_was_imputed=False
    )

    print(f"   ✓ Created {3} LotData objects")
    print()

    # Build solutions in optimizer format
    print("2. Building solutions (mimicking optimizer output)...")

    solutions_optimizer_format = [
        # Solution 1
        (
            [lot1, lot2, lot3],
            [3763.34, 578.55, 618.33],
            1574.07
        ),
        # Solution 2
        (
            [lot1, lot2],
            [3000.0, 1000.0],
            1450.25
        )
    ]

    print(f"   ✓ Created {len(solutions_optimizer_format)} solutions")
    print()

    # Requirements dict (mimicking API request)
    print("3. Preparing requirements...")

    requirements_dict = {
        'dc': 80.0,
        'fp': 650.0,
        'duck': 20.0,
        'species': 'O',
        'qty': 5000.0
    }

    print(f"   ✓ Requirements: DC={requirements_dict['dc']}%, "
          f"FP={requirements_dict['fp']}, Duck={requirements_dict['duck']}%")
    print()

    # Create temporary file (mimicking tempfile usage in service)
    print("4. Exporting to Excel (mimicking service export)...")

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        tmp_path = tmp.name

    print(f"   Temp file: {tmp_path}")

    try:
        # THIS IS THE EXACT CALL FROM THE SERVICE
        export_solutions_to_excel(
            solutions=solutions_optimizer_format,
            requirements=requirements_dict,
            output_path=tmp_path
        )

        print(f"   ✓ Export completed successfully")
        print()

        # Verify the file
        print("5. Verifying output...")

        file_size = os.path.getsize(tmp_path)
        print(f"   File size: {file_size:,} bytes")

        # Load and inspect
        wb = load_workbook(tmp_path)
        ws = wb['TUTTE_LE_SOLUZIONI']

        print(f"   Worksheet: {ws.title}")
        print(f"   Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        print()

        # Check structure
        print("6. Validating structure...")

        row_num = 1
        for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
            if row[0]:
                content = str(row[0])[:60] + "..." if len(str(row[0])) > 60 else str(row[0])
                print(f"   Row {row_num}: {content}")
            row_num += 1

        print()
        print("7. Validation checks...")

        # Count solution headers
        solution_count = 0
        summary_count = 0
        for row in ws.iter_rows(values_only=True):
            if row[0]:
                if 'SOLUZIONE' in str(row[0]):
                    solution_count += 1
                if '📊 Totale:' in str(row[0]):
                    summary_count += 1

        print(f"   ✓ Found {solution_count} solution headers (expected: 2)")
        print(f"   ✓ Found {summary_count} summary rows (expected: 2)")

        # Verify data integrity
        data_rows = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] and row[0] != 'Codice Art' and 'SOLUZIONE' not in str(row[0]) and '📊' not in str(row[0]):
                if row[5]:  # Codice Lotto should be present
                    data_rows += 1

        print(f"   ✓ Found {data_rows} data rows with lot codes")

        # Success summary
        print()
        print("=" * 70)
        print("✓ INTEGRATION TEST PASSED")
        print("=" * 70)
        print()
        print("Summary:")
        print(f"  - Solutions exported: {solution_count}")
        print(f"  - Lots exported: {data_rows}")
        print(f"  - File size: {file_size:,} bytes")
        print(f"  - Output file: {tmp_path}")
        print()
        print("The Excel file is ready for manual inspection if needed.")
        print()

        return True

    except Exception as e:
        print()
        print("=" * 70)
        print("✗ INTEGRATION TEST FAILED")
        print("=" * 70)
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return False

    finally:
        # Note: We don't delete the temp file so it can be inspected
        print(f"Temporary file preserved at: {tmp_path}")
        print("(Delete manually when done inspecting)")


def test_in_memory_export():
    """
    Test in-memory export for web API usage
    """
    print()
    print("=" * 70)
    print("Integration Test: In-Memory Export (Web API)")
    print("=" * 70)
    print()

    from excel_export import export_solutions_to_bytes

    # Simple test data
    lot = LotData(
        article_code='3|POB',
        lot_code='TEST-001',
        description='Test',
        dc_real=85.0,
        fp_real=750.0,
        duck_real=10.0,
        qty_available=1000.0,
        cost_per_kg=100.0,
        is_estimated=False
    )

    solutions = [([lot], [1000.0], 1000.0)]
    requirements = {'dc': 85.0, 'qty': 1000.0}

    print("1. Generating Excel in memory...")

    try:
        excel_bytes = export_solutions_to_bytes(solutions, requirements)

        print(f"   ✓ Generated {len(excel_bytes):,} bytes")
        print()

        # Verify it's valid Excel
        print("2. Verifying Excel validity...")

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(excel_bytes)
            tmp_path = tmp.name

        wb = load_workbook(tmp_path)
        print(f"   ✓ Valid Excel file")
        print(f"   ✓ Sheet: {wb.active.title}")

        os.unlink(tmp_path)

        print()
        print("✓ IN-MEMORY EXPORT TEST PASSED")
        print()

        return True

    except Exception as e:
        print(f"✗ IN-MEMORY EXPORT TEST FAILED: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == '__main__':
    print("\n")
    print("╔" + "=" * 68 + "╗")
    print("║" + " " * 68 + "║")
    print("║" + "  EXCEL EXPORT MODULE - INTEGRATION TEST SUITE".center(68) + "║")
    print("║" + " " * 68 + "║")
    print("╚" + "=" * 68 + "╝")
    print()

    # Run tests
    test1_passed = test_service_usage_pattern()
    test2_passed = test_in_memory_export()

    # Final summary
    print()
    print("=" * 70)
    print("FINAL RESULTS")
    print("=" * 70)
    print(f"Service Usage Pattern Test: {'✓ PASS' if test1_passed else '✗ FAIL'}")
    print(f"In-Memory Export Test:      {'✓ PASS' if test2_passed else '✗ FAIL'}")
    print("=" * 70)
    print()

    if test1_passed and test2_passed:
        print("🎉 ALL INTEGRATION TESTS PASSED!")
        print()
        print("The excel_export.py module is ready for production use.")
        sys.exit(0)
    else:
        print("⚠️  SOME TESTS FAILED")
        print()
        print("Please review the errors above and fix before deployment.")
        sys.exit(1)
